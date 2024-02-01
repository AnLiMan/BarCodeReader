# -----Библиотеки-----
from PyQt6 import QtWidgets, uic, QtGui
from PyQt6.QtGui import QPixmap
from barcode import EAN13 #Есть и другие варианты: EAN-8, EAN-13, EAN-14, UPC-A, JAN, ISBN-10, ISBN-13
from barcode.writer import ImageWriter #Для генерации Штрих-кода
import sys
import os
import openpyxl

#Постоянные
windowName = "BarCodeMainInterface.ui"
picture = "Test.jpg"
excel_filename = 'Данные.xlsx'
groups = ["101", "102", "103", "104", "105"]

#Генерация штрих-кода
def BarCodeGenerate():
    #Создадим папку для сохранения штрих-кодов
    if not os.path.isdir("Codes"):
        os.mkdir("Codes")  # Создание пустого каталога (папки), в которую будут сохраняться фрагменты
        main_window.textBrowser.append("\nКаталог для штрих-кодов создан")

    number = main_window.BarCodeValue.text()
    name = main_window.BarName.text()
    my_code = EAN13(number, writer=ImageWriter())
    my_code.save(f"Codes/{name}")
    main_window.textBrowser.append(f"Для: {number} сгенерирован код с названием: {name}")

    #Структура: [['Имя', 'Фамилия', 'Отчество', 'Штрих-код', 'Имя кода', 'Группа']]
    data_to_write = [[main_window.FirstNameValue.text(), main_window.LastNameValue.text(),main_window.MiddleNameValue.text(),
                      main_window.BarCodeValue.text(), main_window.BarName.text(), groups[main_window.comboBox.currentIndex()]]]

    write_to_excel(excel_filename, data_to_write) #Обновим данные в excel-файле

#Закрыть окно программы+
def CLoseWindow():
    sys.exit(app.exit())

#Отработка Check Box
def Check():
    if main_window.checkGraph.isChecked() == True:
        name = main_window.BarName.text()
        label = main_window.GraphLabel

        #Если существует файл с таким названием
        if os.path.isfile(f"Codes/{name}.png"):
            pixmap = QPixmap(f"Codes/{name}")
        #Иначе изображение-затычка
        else:
            pixmap = QPixmap(picture)
        label.setPixmap(pixmap)
    else:
        main_window.GraphLabel.setText("Здесь будет изображение")

#Вывод справки по программе
def WorkInfo():
    main_window.textBrowser.append("1. Для генерации кода введите фамилию, имя и отчество студента"
                                   " ,а также выберите группу из выпадающего списка, \n2. Нажмите кнопку 'Создать штрихкод'  "
                                   "\n3. Для отображения сгенерированного кода нажмите 'Вкл графику'")
    main_window.textBrowser.append("Исходный код на Git Hub: https://github.com/AnLiMan/BarCodeReader")

#Очистка полей
clearN = lambda: main_window.FirstNameValue.clear()
clearLN = lambda: main_window.LastNameValue.clear()
clearMN = lambda: main_window.MiddleNameValue.clear()
def ClearBarCode():
    main_window.BarCodeValue.clear(),
    main_window.BarName.clear()

#Чтение данных из excel
def read_from_excel(filename):
    try:
        wb = openpyxl.load_workbook(filename, read_only=False)
    except FileNotFoundError:
        # Если файл не существует, создаем новый
        wb = openpyxl.Workbook()

    ws = wb.active

    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)

    return wb, data

#Обновление записей в excel
def write_to_excel(filename, data):
    wb, existing_data = read_from_excel(filename)

    # Очищаем лист перед записью новых данных
    ws = wb.active
    for row in range(1, ws.max_row + 1):
        ws.delete_rows(1, amount=ws.max_row)

    # Добавляем существующие данные
    for row in existing_data:
        ws.append(row)

    # Добавляем новые данные
    for row in data:
        ws.append(row)

    wb.save(filename)


if __name__ == "__main__":
    app = QtWidgets.QApplication([]) #Запускаем приложение
    # Загружаем ранее созданный интерфейс,  здесь и далее  мы будем обращаться к
    #объекту main_window как к нашему основному интерфейсу
    main_window = uic.loadUi(windowName)
    main_window.comboBox.addItems(groups) #Добавим список групп
    main_window.show() #Показываем окно интерфейса

    #Очистка полей
    main_window.ClearName.clicked.connect(clearN)
    main_window.ClearLastName.clicked.connect(clearLN)
    main_window.ClearMiddleName.clicked.connect(clearMN)
    main_window.ClearCodeName.clicked.connect(ClearBarCode)

    main_window.CodeGenerate.clicked.connect(BarCodeGenerate) #Сгенерировать штрих-код
    main_window.CloseWindow.clicked.connect(CLoseWindow) #Закрыть окно
    main_window.CloseWindow2.triggered.connect(CLoseWindow) #Закрыть окно
    main_window.Info.triggered.connect(WorkInfo) #Отобразим вспомогательную информацию
    main_window.checkGraph.clicked.connect(Check)  # Нажали на Check Box

    sys.exit(app.exec())  #Крутим приложение по кругу